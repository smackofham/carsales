""" Contains all of the excel related functions. """
from openpyxl import Workbook


def convert_data_to_excel_readable(list_name, car):
    """ Converts car data into a list of dictionaries.
    Inputs:
    list_name = Name of output list of dictionaries.
    car = Name of object that data is obtained from. """
    """ Name for the self variables related to the car data. """
    # variable_names needs to be in the order of the columns
    variables_name = ['car.year', 'car.name', 'car.odometer', 'car.body',
                      'car.transmission', 'car.engine', 'car.price', 'car.url']
    variables_values = [car.year, car.name, car.odometer, car.body,
                        car.transmission, car.engine, car.price, car.url]
    car_dict_entry = {}
    for i in range(len(variables_name)):
        car_dict_entry[variables_name[i]] = variables_values[i]
    list_name.append(car_dict_entry)


def write_data(name_key, list_dict_car_details):
    """ Opens a new workbook. Sets the name of the sheet/xlsx file as the name of the car.
    """
    wb = Workbook()
    # Allows us to use ws as the reference to the active sheet
    ws = wb.active
    ws.Title = str(list_dict_car_details[0][str(name_key)])

    # Loop to get headings for table.
    # list_dict_car_details is final list of dictionaries of different cars.
    keys = []
    # Builds list of keys so that they can be used to get the dictionary values.
    for k in list_dict_car_details[0]:
        keys.append(k)
    # Iterates through the keys and writes them as column headings
    for i in range(len(list_dict_car_details[0].keys())):
        # Writes the headings in the different columns
        ws.cell(row=1, column=i + 1).value = str(keys[i])
    # Iterates through the list of dictionary entries.
    for j in range(len(list_dict_car_details)):
        # Iterates through the key:value pairs and writes them in the appropriate column.
        for i in range(len(keys)):
            ws.cell(row=j + 2, column=i + 1).value = list_dict_car_details[j][keys[i]]
    print('Tried to save.')
    wb.save("%s.xlsx" % list_dict_car_details[0][str(name_key)])

